from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from typing import List
from io import BytesIO
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .. import models, schemas
from ..database import get_db

router = APIRouter(prefix="/api/students", tags=["students"])


@router.post("/kindergartens/", response_model=schemas.Kindergarten)
def create_kindergarten(kindergarten: schemas.KindergartenCreate, db: Session = Depends(get_db)):
    db_kindergarten = models.Kindergarten(**kindergarten.model_dump())
    db.add(db_kindergarten)
    db.commit()
    db.refresh(db_kindergarten)
    return db_kindergarten


@router.get("/kindergartens/", response_model=List[schemas.Kindergarten])
def get_kindergartens(db: Session = Depends(get_db)):
    return db.query(models.Kindergarten).all()


@router.get("/kindergartens/{kindergarten_id}", response_model=schemas.Kindergarten)
def get_kindergarten(kindergarten_id: int, db: Session = Depends(get_db)):
    kindergarten = db.query(models.Kindergarten).filter(models.Kindergarten.id == kindergarten_id).first()
    if not kindergarten:
        raise HTTPException(status_code=404, detail="Kindergarten not found")
    return kindergarten


@router.post("/classes/", response_model=schemas.KindergartenClass)
def create_class(class_data: schemas.ClassCreate, db: Session = Depends(get_db)):
    kindergarten = db.query(models.Kindergarten).filter(
        models.Kindergarten.id == class_data.kindergarten_id
    ).first()
    if not kindergarten:
        raise HTTPException(status_code=404, detail="Kindergarten not found")

    db_class = models.KindergartenClass(**class_data.model_dump())
    db.add(db_class)
    db.commit()
    db.refresh(db_class)
    return db_class


@router.get("/classes/", response_model=List[schemas.KindergartenClass])
def get_classes(kindergarten_id: int = None, db: Session = Depends(get_db)):
    query = db.query(models.KindergartenClass)
    if kindergarten_id:
        query = query.filter(models.KindergartenClass.kindergarten_id == kindergarten_id)
    return query.all()


@router.get("/classes/{class_id}", response_model=schemas.KindergartenClass)
def get_class(class_id: int, db: Session = Depends(get_db)):
    class_ = db.query(models.KindergartenClass).filter(models.KindergartenClass.id == class_id).first()
    if not class_:
        raise HTTPException(status_code=404, detail="Class not found")
    return class_


@router.post("/", response_model=schemas.Student)
def create_student(student: schemas.StudentCreate, db: Session = Depends(get_db)):
    class_ = db.query(models.KindergartenClass).filter(
        models.KindergartenClass.id == student.class_id
    ).first()
    if not class_:
        raise HTTPException(status_code=404, detail="Class not found")

    db_student = models.Student(**student.model_dump())
    db.add(db_student)
    db.commit()
    db.refresh(db_student)
    return db_student


@router.get("/", response_model=List[schemas.Student])
def get_students(class_id: int = None, db: Session = Depends(get_db)):
    query = db.query(models.Student)
    if class_id:
        query = query.filter(models.Student.class_id == class_id)
    return query.all()


@router.get("/template/download")
def download_template():
    """학생 일괄 등록용 엑셀 템플릿 다운로드"""
    wb = Workbook()
    ws = wb.active
    ws.title = "학생등록"

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 헤더 작성
    headers = ["유치원명", "유치원주소", "반이름", "학생이름", "나이"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 예시 데이터
    example_data = [
        ["햇살유치원", "서울시 강남구 테헤란로 123", "햇님반", "김민준", 5],
        ["햇살유치원", "서울시 강남구 테헤란로 123", "햇님반", "이서연", 5],
        ["햇살유치원", "서울시 강남구 테헤란로 123", "달님반", "박지호", 6],
        ["꿈나무유치원", "서울시 서초구 반포대로 456", "별님반", "최수아", 4],
    ]

    for row_idx, row_data in enumerate(example_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 컬럼 너비 조정
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10

    # 파일 저장
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=student_template.xlsx"}
    )


@router.post("/upload/excel")
def upload_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """엑셀 파일로 학생 일괄 등록"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="엑셀 파일(.xlsx, .xls)만 업로드 가능합니다.")

    try:
        contents = file.file.read()
        wb = load_workbook(filename=BytesIO(contents))
        ws = wb.active

        created_kindergartens = {}
        created_classes = {}
        created_students = []
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            if not any(row):  # 빈 행 스킵
                continue

            kg_name, kg_address, class_name, student_name, age = row[:5]

            # 유효성 검사
            if not all([kg_name, class_name, student_name]):
                errors.append(f"행 {row_idx}: 필수 항목(유치원명, 반이름, 학생이름)이 누락되었습니다.")
                continue

            try:
                age = int(age) if age else 5
            except (ValueError, TypeError):
                age = 5

            # 유치원 생성 또는 조회
            kg_key = str(kg_name).strip()
            if kg_key not in created_kindergartens:
                existing_kg = db.query(models.Kindergarten).filter(
                    models.Kindergarten.name == kg_key
                ).first()

                if existing_kg:
                    created_kindergartens[kg_key] = existing_kg
                else:
                    new_kg = models.Kindergarten(
                        name=kg_key,
                        address=str(kg_address).strip() if kg_address else "",
                        region=""
                    )
                    db.add(new_kg)
                    db.flush()
                    created_kindergartens[kg_key] = new_kg

            kindergarten = created_kindergartens[kg_key]

            # 반 생성 또는 조회
            class_key = f"{kg_key}_{str(class_name).strip()}"
            if class_key not in created_classes:
                existing_class = db.query(models.KindergartenClass).filter(
                    models.KindergartenClass.kindergarten_id == kindergarten.id,
                    models.KindergartenClass.name == str(class_name).strip()
                ).first()

                if existing_class:
                    created_classes[class_key] = existing_class
                else:
                    new_class = models.KindergartenClass(
                        name=str(class_name).strip(),
                        kindergarten_id=kindergarten.id
                    )
                    db.add(new_class)
                    db.flush()
                    created_classes[class_key] = new_class

            class_ = created_classes[class_key]

            # 학생 생성
            new_student = models.Student(
                name=str(student_name).strip(),
                age=age,
                class_id=class_.id
            )
            db.add(new_student)
            created_students.append(str(student_name).strip())

        db.commit()

        return {
            "message": "업로드 완료",
            "created_kindergartens": len(set(created_kindergartens.keys())),
            "created_classes": len(set(created_classes.keys())),
            "created_students": len(created_students),
            "errors": errors
        }

    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"파일 처리 중 오류 발생: {str(e)}")


@router.get("/{student_id}", response_model=schemas.Student)
def get_student(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")
    return student


@router.put("/{student_id}", response_model=schemas.Student)
def update_student(student_id: int, student_update: schemas.StudentUpdate, db: Session = Depends(get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    update_data = student_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(student, key, value)

    db.commit()
    db.refresh(student)
    return student


@router.delete("/{student_id}")
def delete_student(student_id: int, db: Session = Depends(get_db)):
    student = db.query(models.Student).filter(models.Student.id == student_id).first()
    if not student:
        raise HTTPException(status_code=404, detail="Student not found")

    db.delete(student)
    db.commit()
    return {"message": "Student deleted successfully"}


@router.delete("/kindergartens/{kindergarten_id}")
def delete_kindergarten(kindergarten_id: int, db: Session = Depends(get_db)):
    kindergarten = db.query(models.Kindergarten).filter(models.Kindergarten.id == kindergarten_id).first()
    if not kindergarten:
        raise HTTPException(status_code=404, detail="Kindergarten not found")

    db.delete(kindergarten)
    db.commit()
    return {"message": "Kindergarten deleted successfully"}


@router.delete("/classes/{class_id}")
def delete_class(class_id: int, db: Session = Depends(get_db)):
    class_ = db.query(models.KindergartenClass).filter(models.KindergartenClass.id == class_id).first()
    if not class_:
        raise HTTPException(status_code=404, detail="Class not found")

    db.delete(class_)
    db.commit()
    return {"message": "Class deleted successfully"}
